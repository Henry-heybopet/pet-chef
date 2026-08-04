#!/usr/bin/env python3
"""Import the recipe Excel workbook into backend/src/data/recipes_db.js.

The Excel workbook is the human-edited recipe source. This script updates the
rawRecipes array only, preserving the existing computed fields and helpers in
recipes_db.js.
"""

from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Missing dependency: install openpyxl for python3 before importing recipes.") from exc


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_EXCEL = REPO_ROOT / "docs/source/犬用鲜食配方_A+B_40种优化版_营养合规审查（0630）.xlsx"
RECIPES_JS = REPO_ROOT / "backend/src/data/recipes_db.js"
RECIPE_SHEET = "01_40个食谱_A+B"
RAW_START = "const rawRecipes = ["
RAW_END = "];\n\n// 计算每个食谱的含水量、功效说明、烹饪基准参数"

CATEGORY_DEFAULTS = {
    "幼犬通用": {"category_code": 1, "category_type": "life_stage_size", "life_stage": "幼犬", "dog_size": None},
    "控钙幼犬（大型幼犬）": {"category_code": 3, "category_type": "life_stage_size", "life_stage": "幼犬", "dog_size": "大型犬"},
    "成犬通用": {"category_code": 4, "category_type": "life_stage_size", "life_stage": "成年犬", "dog_size": None},
    "老年犬通用": {"category_code": 7, "category_type": "life_stage_size", "life_stage": "老年犬", "dog_size": None},
    "美毛护肤": {"category_code": 10, "category_type": "functional", "life_stage": None, "dog_size": None},
    "护肝": {"category_code": 11, "category_type": "functional", "life_stage": None, "dog_size": None},
    "低敏单一蛋白": {"category_code": 14, "category_type": "functional", "life_stage": None, "dog_size": None},
}
INGREDIENT_ALIASES = {
    "西蓝花": "西兰花",
}


def existing_meta() -> dict[str, dict]:
    script = """
const { recipesDb } = require('./backend/src/data/recipes_db');
console.log(JSON.stringify(Object.fromEntries(recipesDb.map(r => [r.id, {
  category_code: r.category_code,
  category_type: r.category_type,
  life_stage: r.life_stage ?? null,
  dog_size: r.dog_size ?? null,
  tags: r.tags || []
}]))));
"""
    output = subprocess.check_output(["node", "-e", script], cwd=REPO_ROOT, text=True)
    return json.loads(output)


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_percent(value) -> float:
    number = float(value)
    return int(number) if number.is_integer() else round(number, 3)


def parse_ingredients(value: str) -> dict[str, float]:
    ingredients: dict[str, float] = {}
    for part in normalize_text(value).split("/"):
        text = part.strip()
        if not text:
            continue
        match = re.match(r"^(.+?)\s*([0-9]+(?:\.[0-9]+)?)\s*%?$", text)
        if not match:
            raise ValueError(f"Cannot parse ingredient item: {text}")
        name = re.sub(r"\s+", "", match.group(1).strip())
        name = INGREDIENT_ALIASES.get(name, name)
        ingredients[name] = parse_percent(match.group(2))
    if not ingredients:
        raise ValueError("A包明细 is empty")
    return ingredients


def import_rows(excel_path: Path) -> list[dict]:
    workbook = load_workbook(excel_path, data_only=False)
    if RECIPE_SHEET not in workbook.sheetnames:
        raise ValueError(f"Missing sheet: {RECIPE_SHEET}")
    sheet = workbook[RECIPE_SHEET]
    headers = [normalize_text(cell.value) for cell in sheet[1]]
    column = {name: index for index, name in enumerate(headers)}
    required = ["recipe_id", "产品化大类", "新版配方名称", "A包明细", "B包明细"]
    missing = [name for name in required if name not in column]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    meta = existing_meta()
    rows: list[dict] = []
    seen: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        recipe_id = normalize_text(row[column["recipe_id"]])
        if not recipe_id:
            continue
        if recipe_id in seen:
            raise ValueError(f"Duplicate recipe_id: {recipe_id}")
        seen.add(recipe_id)

        category = normalize_text(row[column["产品化大类"]])
        name = normalize_text(row[column["新版配方名称"]])
        if not category or not name:
            raise ValueError(f"{recipe_id}: category/name is required")

        defaults = CATEGORY_DEFAULTS.get(category, {})
        old = meta.get(recipe_id, {})
        rows.append({
            "id": recipe_id,
            "category": category,
            "category_code": old.get("category_code", defaults.get("category_code")),
            "category_type": old.get("category_type", defaults.get("category_type")),
            "life_stage": old.get("life_stage", defaults.get("life_stage")),
            "dog_size": old.get("dog_size", defaults.get("dog_size")),
            "name": name,
            "tags": old.get("tags", []),
            "ingredients": parse_ingredients(row[column["A包明细"]]),
            "b_pack": normalize_text(row[column["B包明细"]]),
        })

    if len(rows) != 40:
        raise ValueError(f"Expected 40 recipes, got {len(rows)}")
    return rows


def write_recipes_js(rows: list[dict]) -> None:
    text = RECIPES_JS.read_text(encoding="utf-8")
    start = text.index(RAW_START)
    end = text.index(RAW_END)
    raw_json = json.dumps(rows, ensure_ascii=False, indent=2)
    next_text = text[:start] + f"const rawRecipes = {raw_json};" + text[end + len("];"):]
    RECIPES_JS.write_text(next_text, encoding="utf-8")


def validate_generated_js() -> None:
    script = """
delete require.cache[require.resolve('./backend/src/data/recipes_db')];
const { recipesDb } = require('./backend/src/data/recipes_db');
const allowedProteinGroups = new Set(['chicken', 'duck', 'beef', 'fish', 'rabbit', 'other']);
const missing = recipesDb.filter(r => {
  const c = r.cooking_base || {};
  return !c.temperature || !c.power || !c.speed || !c.cook_minutes || !c.preheat_minutes || !c.texture_profile || !allowedProteinGroups.has(c.protein_group);
}).map(r => r.id);
if (recipesDb.length !== 40 || missing.length) {
  console.error(JSON.stringify({ count: recipesDb.length, missing }, null, 2));
  process.exit(1);
}
console.log(JSON.stringify({
  count: recipesDb.length,
  cookingSamples: recipesDb.slice(0, 5).map(r => ({
    id: r.id,
    name: r.name,
    cooking_base: r.cooking_base,
  })),
}, null, 2));
"""
    subprocess.check_call(["node", "-e", script], cwd=REPO_ROOT)


def main() -> None:
    excel_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else DEFAULT_EXCEL
    if not excel_path.is_absolute():
        excel_path = (REPO_ROOT / excel_path).resolve()
    if not excel_path.exists():
        raise SystemExit(f"Excel file not found: {excel_path}")
    rows = import_rows(excel_path)
    write_recipes_js(rows)
    validate_generated_js()
    print(f"Imported {len(rows)} recipes from {excel_path}")
    print(f"Updated {RECIPES_JS}")


if __name__ == "__main__":
    main()
